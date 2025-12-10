from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
import datetime 
from utils.FileStorage import upload_file

import random
from datetime import date
# Load an existing workbook

def formatDate(datestring):
    datestr = str(datestring).split("-")
    
    return f"{datestr[2]}-{datestr[1]}-{datestr[0]}"

def prepareCells(workshop, start, end):
    workshop.merge_cells(f"{start}:{end}")
    
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in workshop[f"{start}:{end}"]:
        for cell in row:
            cell.border = border_style


def createNewOPCR(data, assigned, admin_data):
    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    period = ""
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year
        
    print("ADMIN DATA", admin_data)
    print("OPCR_DATA", data)

    wb = load_workbook("excels/OPCRTest.xlsx")
    ws = wb.active
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    name = admin_data["fullName"]
    ws["A6"] = f"I, {name}, {admin_data['position']} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = admin_data["position"]
    ws["N10"] = ""
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")
    
    ws["A13"] = "Name: " + admin_data["individuals"]["approve"]["name"]
    ws["A14"] = "Position: " + admin_data["individuals"]["approve"]["position"]
    ws["A15"] = "Date: " + ""
    
    row = 25
    startingrow = 25
    endrow = 0
    numberformat = "0.00"

    # Iterate through function types: [{"Core Function": {...}}, {"Support Function": {...}}, ...]
    for func_type, func_type_dict in data.items():

        print("OPCR DATA",func_type_dict)
        if len(func_type_dict) == 0:
            continue

        ws[f"A{row}"] = func_type
        ws[f"A{row}"].font = Font(bold=True, size=12)
        prepareCells(ws, f"A{row}", f"S{row}")
        row += 1
        
        # Extract function type and its categories
        for cat_name, categories_dict in func_type_dict.items():
            if len(categories_dict) == 0:
                continue
            
            # Write function type header
            ws[f"A{row}"] = cat_name
            ws[f"A{row}"].font = Font(bold=True, size=12)
            prepareCells(ws, f"A{row}", f"S{row}")
            row += 1

            print("CAT DICT: ",categories_dict)
            
            # Iterate through categories within this function type
            for a in categories_dict:
                if len(a) == 0:
                    continue
                
                prepareCells(ws, f"A{row}", f"C{row+5}")
                ws[f"A{row}"] = a["title"]
                ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Prepare target cells
                prepareCells(ws, f"D{row}", f"D{row+1}")
                ws[f"D{row}"] = a["summary"]["target"]
                ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"D{row}"].font = Font(bold=True)

                if a["description"]["timeliness_mode"] == "deadline":
                    prepareCells(ws, f"D{row+2}", f"D{row+3}")
                    ws[f"D{row+2}"] = ""
                    ws[f"D{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"D{row+2}"].font = Font(bold=True)
                else:
                    prepareCells(ws, f"D{row+2}", f"D{row+3}")
                    ws[f"D{row+2}"] = a["working_days"]["target"]
                    ws[f"D{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"D{row+2}"].font = Font(bold=True)
                    
                prepareCells(ws, f"D{row+4}", f"D{row+5}")
                ws[f"D{row+4}"] = a["corrections"]["target"]
                ws[f"D{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"D{row+4}"].font = Font(bold=True)
                    
                    # Description
                prepareCells(ws, f"E{row}", f"G{row+1}")
                ws[f"E{row}"] = a["description"]["target"]
                ws[f"E{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                ws.row_dimensions[row].height = 45

                if a["description"]["timeliness_mode"] == "deadline":
                    prepareCells(ws, f"E{row+2}", f"G{row+3}")
                    ws[f"E{row+2}"] = "on the set deadline"
                    ws[f"E{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    prepareCells(ws, f"E{row+2}", f"G{row+3}")
                    ws[f"E{row+2}"] = str(a["description"]["time"] or "") + " spent"
                    ws[f"E{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                prepareCells(ws, f"E{row+4}", f"G{row+5}")
                ws[f"E{row+4}"] = a["description"]["alterations"]
                ws[f"E{row+4}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
                    # Budget allotted
                prepareCells(ws, f"H{row}", f"H{row+5}")
                ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"H{row}"] = 0
                    
                    # Individuals involved
                accountable = ""
                for person in assigned.get(a["title"], []):
                    accountable += person + "\n"
                    
                prepareCells(ws, f"I{row}", f"I{row+5}")
                ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"I{row}"] = accountable
                    
                    # Actual cells
                prepareCells(ws, f"J{row}", f"J{row+1}")
                ws[f"J{row}"] = a["summary"]["actual"]
                ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"J{row}"].font = Font(bold=True)

                if a["description"]["timeliness_mode"] == "deadline":
                    prepareCells(ws, f"J{row+2}", f"J{row+3}")
                    ws[f"J{row+2}"] = abs(int(a["working_days"]["actual"] / a.get("frequency", 1)))
                    ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+2}"].font = Font(bold=True)
                else:
                    prepareCells(ws, f"J{row+2}", f"J{row+3}")
                    ws[f"J{row+2}"] = a["working_days"]["actual"]
                    ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+2}"].font = Font(bold=True)
                    
                prepareCells(ws, f"J{row+4}", f"J{row+5}")
                ws[f"J{row+4}"] = int(a["corrections"]["actual"] / a.get("frequency", 1))
                ws[f"J{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"J{row+4}"].font = Font(bold=True)
                    
                prepareCells(ws, f"K{row}", f"M{row+1}")
                ws[f"K{row}"] = a["description"]["actual"]
                ws[f"K{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                if a["description"]["timeliness_mode"] == "deadline":
                    actual_deadline_desc = "on the set deadline"
                    if a["working_days"]["actual"] == 0:
                        actual_deadline_desc = "on the set deadline"
                    elif a["working_days"]["actual"] > 0:
                        actual_deadline_desc = "day/s late in average"
                    else:
                        actual_deadline_desc = "day/s early in average"

                    prepareCells(ws, f"K{row+2}", f"M{row+3}")
                    ws[f"K{row+2}"] = actual_deadline_desc
                    ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                else:
                    prepareCells(ws, f"K{row+2}", f"M{row+3}")
                    ws[f"K{row+2}"] = str(a["description"]["time"] or "") + " spent"
                    ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    
                prepareCells(ws, f"K{row+4}", f"M{row+5}")
                ws[f"K{row+4}"] = a["description"]["alterations"] + "/s in average"
                ws[f"K{row+4}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    
                    # Ratings
                prepareCells(ws, f"N{row}", f"N{row+5}")
                ws[f"N{row}"] = a["rating"]["quantity"]
                ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"N{row}"].font = Font(bold=True)
                    
                prepareCells(ws, f"O{row}", f"O{row+5}")
                ws[f"O{row}"] = a["rating"]["efficiency"]
                ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"O{row}"].font = Font(bold=True)
                    
                prepareCells(ws, f"P{row}", f"P{row+5}")
                ws[f"P{row}"] = a["rating"]["timeliness"]
                ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"P{row}"].font = Font(bold=True)
                    
                prepareCells(ws, f"Q{row}", f"Q{row+5}")
                ws[f"Q{row}"] = a["rating"]["average"]
                ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"Q{row}"].font = Font(bold=True)
                    
                    # Remarks with dynamic formula
                from models.System_Settings import System_Settings
                settings = System_Settings.query.first()
                rating_thresholds = settings.rating_thresholds if settings else {
                        "POOR": {"min": 1, "max": 1.9},
                        "UNSATISFACTORY": {"min": 2, "max": 2.9},
                        "SATISFACTORY": {"min": 3, "max": 3.9},
                        "VERY SATISFACTORY": {"min": 4, "max": 4.9},
                        "OUTSTANDING": {"min": 5, "max": 5}
                    }

                average = a["rating"]["average"]
                formula_parts = []
                for rating_name, thresholds in rating_thresholds.items():
                    min_val = thresholds.get("min", 0)
                    max_val = thresholds.get("max", 5)
                    if min_val == max_val:
                        formula_parts.append(f'IF({average}={min_val}, "{str(rating_name).replace("_", " ").upper()}"')
                    else:
                        formula_parts.append(f'IF(AND({average}>={min_val}, {average}<={max_val}), "{str(rating_name).replace("_", " ").upper()}"')
                    
                formula = ""
                for i, part in enumerate(formula_parts):
                    formula += part + ", "
                formula += '"UNRATED"' + ")" * len(formula_parts)

                prepareCells(ws, f"R{row}", f"S{row+5}")
                ws[f"R{row}"] = f"={formula}"
                ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws[f"R{row}"].font = Font(bold=True)
                
                endrow = row
                row = row + 6

    # Final averages and adjectival rating
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "Final Average Rating"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"N{row+1}")
    ws[f"N{row}"] = f"=AVERAGE(N{startingrow+1}:N{endrow})"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].number_format = numberformat
    
    prepareCells(ws, f"O{row}", f"O{row+1}")
    ws[f"O{row}"] = f"=AVERAGE(O{startingrow+1}:O{endrow})"
    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"O{row}"].number_format = numberformat
    
    prepareCells(ws, f"P{row}", f"P{row+1}")
    ws[f"P{row}"] = f"=AVERAGE(P{startingrow+1}:P{endrow})"
    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"P{row}"].number_format = numberformat
    
    average = f"=AVERAGE(Q{startingrow}:Q{endrow})"
    averagecell = f"Q{row}"
    prepareCells(ws, f"Q{row}", f"Q{row+1}")
    ws[f"Q{row}"] = f"=AVERAGE(Q{startingrow+1}:Q{endrow})"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].number_format = numberformat
    
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "FINAL AVERAGE RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = average
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)
    ws[f"N{row}"].number_format = numberformat    
    
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "ADJECTIVAL RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = f'=IF(AND({averagecell}>=1, {averagecell}<=1.9), "POOR", IF(AND({averagecell}>=2, {averagecell}<=2.9), "UNSATISFACTORY", IF(AND({averagecell}>=3, {averagecell}<=3.9), "SATISFACTORY", IF(AND({averagecell}>=4, {averagecell}<=4.9), "VERY SATISFACTORY", IF(AND({averagecell}=5), "OUTSTANDING")))))'
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)  

    
    # Individuals section
    row += 3
    prepareCells(ws, f"A{row}", f"D{row}")
    ws["A"+str(row)] = "Discussed With"
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"E{row}", f"F{row}")
    ws["E"+str(row)] = "Date"
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"G{row}", f"J{row}")
    ws["G"+str(row)] = "Assessed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"K{row}", f"L{row}")
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"M{row}", f"P{row}")
    ws["M"+str(row)] = "Final Rating By"
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"Q{row}", f"S{row}")
    ws["Q"+str(row)] = "Date"
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row += 1
    
    # Adjust row height for individuals
    # Name + position
    prepareCells(ws, f"A{row}", f"D{row+2}")
    ws["A"+str(row)] = admin_data["individuals"]["discuss"]["name"] + "\n" + admin_data["individuals"]["discuss"]["position"]
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"E{row}", f"F{row+2}")
    ws["E"+str(row)] = ""
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["E"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws["G"+str(row)] = admin_data["individuals"]["assess"]["name"] + "\n" + admin_data["individuals"]["assess"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"M{row}", f"P{row+2}")
    ws["M"+str(row)] = admin_data["individuals"]["final"]["name"] + "\n" + admin_data["individuals"]["final"]["position"]
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"Q{row}", f"S{row+2}")
    ws["Q"+str(row)] = ""
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["Q"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    row+= 4
    
    prepareCells(ws, f"G{row}", f"J{row}")
    ws["G"+str(row)] = "Confirmed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"K{row}", f"L{row}")
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row += 1
    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws["G"+str(row)] = admin_data["individuals"]["confirm"]["name"] + "\n" + admin_data["individuals"]["confirm"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
            
    #prepare category name    
    # Save changes
    id = random.randint(1,999999)
    filename = f"OPCR-NC-{period}-{admin_data['givenName']}-{admin_data['middleName']}-{admin_data['lastName']}-{datee}-{id}"
    link = f"excels/OPCR/{filename}.xlsx"
    wb.save(link)

    file_url = upload_file(link, "commithub-bucket", f"OPCR/{filename}.xlsx")
    
    return file_url


def createNewMasterOPCR(data, assigned, admin_data):
    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    period = ""
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year
        
    print("MASTER OPCR: ", data)
    wb = load_workbook("excels/OPCRTest.xlsx")
    ws = wb.active
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    name = admin_data["fullName"]
    ws["A6"] = f"I, {name}, {admin_data['position']} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = admin_data["position"]
    ws["N10"] = ""
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")
    
    ws["A13"] = "Name: " + admin_data["individuals"]["approve"]["name"]
    ws["A14"] = "Position: " + admin_data["individuals"]["approve"]["position"]
    ws["A15"] = "Date: " + ""
    
    row = 25
    startingrow = 25
    endrow = 0
    numberformat = "0.00"

    # Iterate through function types: [{"Core Function": {...}}, {"Support Function": {...}}, ...]
    for func_type_dict in data:
        if not isinstance(func_type_dict, dict) or len(func_type_dict) == 0:
            continue
        
        # Extract function type and its categories
        for func_type, categories_dict in func_type_dict.items():


            if not isinstance(categories_dict, dict) or len(categories_dict) == 0:
                continue
            
            # Write function type header
            ws[f"A{row}"] = func_type
            ws[f"A{row}"].font = Font(bold=True, size=12)
            prepareCells(ws, f"A{row}", f"S{row}")
            row += 1
            
            # Iterate through categories within this function type
            for cat_name, tasks_list in categories_dict.items():
                if not isinstance(tasks_list, list) or len(tasks_list) == 0:
                    continue

                
                
                # Write category name
                ws[f"A{row}"] = cat_name
                ws[f"A{row}"].font = Font(bold=True, size=11)
                prepareCells(ws, f"A{row}", f"S{row}")
                row += 1
                
                # Iterate through tasks in this category
                for a in tasks_list:
                    print("TASK: ", a)
                    # Prepare output cell
                    prepareCells(ws, f"A{row}", f"C{row+5}")
                    ws[f"A{row}"] = a["title"]
                    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    # Prepare target cells
                    prepareCells(ws, f"D{row}", f"D{row+1}")
                    ws[f"D{row}"] = a["summary"]["target"]
                    ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"D{row}"].font = Font(bold=True)

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"D{row+2}", f"D{row+3}")
                        ws[f"D{row+2}"] = ""
                        ws[f"D{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"D{row+2}"].font = Font(bold=True)
                    else:
                        prepareCells(ws, f"D{row+2}", f"D{row+3}")
                        ws[f"D{row+2}"] = a["working_days"]["target"]
                        ws[f"D{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"D{row+2}"].font = Font(bold=True)
                    
                    prepareCells(ws, f"D{row+4}", f"D{row+5}")
                    ws[f"D{row+4}"] = a["corrections"]["target"]
                    ws[f"D{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"D{row+4}"].font = Font(bold=True)
                        
                        # Description
                    prepareCells(ws, f"E{row}", f"G{row+1}")
                    ws[f"E{row}"] = a["description"]["target"]
                    ws[f"E{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws.row_dimensions[row].height = 45

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"E{row+2}", f"G{row+3}")
                        ws[f"E{row+2}"] = "on the set deadline"
                        ws[f"E{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        prepareCells(ws, f"E{row+2}", f"G{row+3}")
                        ws[f"E{row+2}"] = str(a["description"]["time"] or "") + " spent"
                        ws[f"E{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    prepareCells(ws, f"E{row+4}", f"G{row+5}")
                    ws[f"E{row+4}"] = a["description"]["alterations"]
                    ws[f"E{row+4}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # Budget allotted
                    prepareCells(ws, f"H{row}", f"H{row+5}")
                    ws[f"H{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"H{row}"] = 0
                        
                        # Individuals involved
                    accountable = ""
                    for person in assigned.get(a["title"], []):
                        accountable += person + "\n"
                        
                    prepareCells(ws, f"I{row}", f"I{row+5}")
                    ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"I{row}"] = accountable
                        
                        # Actual cells
                    prepareCells(ws, f"J{row}", f"J{row+1}")
                    ws[f"J{row}"] = a["summary"]["actual"]
                    ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row}"].font = Font(bold=True)

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = abs(int(a["working_days"]["actual"] / a.get("frequency", 1)))
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)
                    else:
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)
                        
                    prepareCells(ws, f"J{row+4}", f"J{row+5}")
                    ws[f"J{row+4}"] = int(a["corrections"]["actual"] / a.get("frequency", 1))
                    ws[f"J{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+4}"].font = Font(bold=True)
                        
                    prepareCells(ws, f"K{row}", f"M{row+1}")
                    ws[f"K{row}"] = a["description"]["actual"]
                    ws[f"K{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    if a["description"]["timeliness_mode"] == "deadline":
                        actual_deadline_desc = "on the set deadline"
                        if a["working_days"]["actual"] == 0:
                            actual_deadline_desc = "on the set deadline"
                        elif a["working_days"]["actual"] > 0:
                            actual_deadline_desc = "day/s late in average"
                        else:
                            actual_deadline_desc = "day/s early in average"

                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = actual_deadline_desc
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    else:
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = str(a["description"]["time"] or "") + " spent"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                        
                    prepareCells(ws, f"K{row+4}", f"M{row+5}")
                    ws[f"K{row+4}"] = a["description"]["alterations"] + "/s in average"
                    ws[f"K{row+4}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                        
                        # Ratings
                    prepareCells(ws, f"N{row}", f"N{row+5}")
                    ws[f"N{row}"] = a["rating"]["quantity"]
                    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"N{row}"].font = Font(bold=True)
                        
                    prepareCells(ws, f"O{row}", f"O{row+5}")
                    ws[f"O{row}"] = a["rating"]["efficiency"]
                    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"O{row}"].font = Font(bold=True)
                        
                    prepareCells(ws, f"P{row}", f"P{row+5}")
                    ws[f"P{row}"] = a["rating"]["timeliness"]
                    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"P{row}"].font = Font(bold=True)
                        
                    prepareCells(ws, f"Q{row}", f"Q{row+5}")
                    ws[f"Q{row}"] = a["rating"]["average"]
                    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"Q{row}"].font = Font(bold=True)
                        
                        # Remarks with dynamic formula
                    from models.System_Settings import System_Settings
                    settings = System_Settings.query.first()
                    rating_thresholds = settings.rating_thresholds if settings else {
                            "POOR": {"min": 1, "max": 1.9},
                            "UNSATISFACTORY": {"min": 2, "max": 2.9},
                            "SATISFACTORY": {"min": 3, "max": 3.9},
                            "VERY SATISFACTORY": {"min": 4, "max": 4.9},
                            "OUTSTANDING": {"min": 5, "max": 5}
                        }

                    average = a["rating"]["average"]
                    formula_parts = []
                    for rating_name, thresholds in rating_thresholds.items():
                        min_val = thresholds.get("min", 0)
                        max_val = thresholds.get("max", 5)
                        if min_val == max_val:
                            formula_parts.append(f'IF({average}={min_val}, "{str(rating_name).replace("_", " ").upper()}"')
                        else:
                            formula_parts.append(f'IF(AND({average}>={min_val}, {average}<={max_val}), "{str(rating_name).replace("_", " ").upper()}"')
                        
                    formula = ""
                    for i, part in enumerate(formula_parts):
                        formula += part + ", "
                    formula += '"UNRATED"' + ")" * len(formula_parts)

                    prepareCells(ws, f"R{row}", f"S{row+5}")
                    ws[f"R{row}"] = f"={formula}"
                    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"R{row}"].font = Font(bold=True)
                    
                    endrow = row
                    row = row + 6

    # Final averages and adjectival rating
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "Final Average Rating"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"N{row+1}")
    ws[f"N{row}"] = f"=AVERAGE(N{startingrow+1}:N{endrow})"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].number_format = numberformat
    
    prepareCells(ws, f"O{row}", f"O{row+1}")
    ws[f"O{row}"] = f"=AVERAGE(O{startingrow+1}:O{endrow})"
    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"O{row}"].number_format = numberformat
    
    prepareCells(ws, f"P{row}", f"P{row+1}")
    ws[f"P{row}"] = f"=AVERAGE(P{startingrow+1}:P{endrow})"
    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"P{row}"].number_format = numberformat
    
    average = f"=AVERAGE(Q{startingrow}:Q{endrow})"
    averagecell = f"Q{row}"
    prepareCells(ws, f"Q{row}", f"Q{row+1}")
    ws[f"Q{row}"] = f"=AVERAGE(Q{startingrow+1}:Q{endrow})"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].number_format = numberformat
    
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "FINAL AVERAGE RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = average
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)
    ws[f"N{row}"].number_format = numberformat    
    
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "ADJECTIVAL RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = f'=IF(AND({averagecell}>=1, {averagecell}<=1.9), "POOR", IF(AND({averagecell}>=2, {averagecell}<=2.9), "UNSATISFACTORY", IF(AND({averagecell}>=3, {averagecell}<=3.9), "SATISFACTORY", IF(AND({averagecell}>=4, {averagecell}<=4.9), "VERY SATISFACTORY", IF(AND({averagecell}=5), "OUTSTANDING")))))'
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)  

    
    # Individuals section
    row += 3
    prepareCells(ws, f"A{row}", f"D{row}")
    ws["A"+str(row)] = "Discussed With"
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"E{row}", f"F{row}")
    ws["E"+str(row)] = "Date"
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"G{row}", f"J{row}")
    ws["G"+str(row)] = "Assessed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"K{row}", f"L{row}")
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, f"M{row}", f"P{row}")
    ws["M"+str(row)] = "Final Rating By"
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, f"Q{row}", f"S{row}")
    ws["Q"+str(row)] = "Date"
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row += 1
    
    #adjust row height for individuals
    # name + position
    prepareCells(ws, str("A"+str(row)), str("D"+str(row+2)))
    ws["A"+str(row)] = admin_data["individuals"]["discuss"]["name"] + "\n" + admin_data["individuals"]["discuss"]["position"]
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("E"+str(row)), str("F"+str(row+2)))
    ws["E"+str(row)] = ""
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["E"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("G"+str(row)), str("J"+str(row+2)))
    ws["G"+str(row)] = admin_data["individuals"]["assess"]["name"] + "\n" + admin_data["individuals"]["assess"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row+2)))
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("M"+str(row)), str("P"+str(row+2)))
    ws["M"+str(row)] = admin_data["individuals"]["final"]["name"] + "\n" + admin_data["individuals"]["final"]["position"]
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("Q"+str(row)), str("S"+str(row+2)))
    ws["Q"+str(row)] = ""
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["Q"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    row+= 4
    
    prepareCells(ws, str("G"+str(row)), str("J"+str(row)))
    ws["G"+str(row)] = "Confirmed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row)))
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row += 1
    prepareCells(ws, str("G"+str(row)), str("J"+str(row+2)))
    ws["G"+str(row)] = admin_data["individuals"]["confirm"]["name"] + "\n" + admin_data["individuals"]["confirm"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row+2)))
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    # Save changes
    id = random.randint(1, 999999)
    filename = f"Master-OPCR-NC-{period}-{admin_data['givenName']}-{admin_data['middleName']}-{admin_data['lastName']}-{datee}-{id}"
    link = f"excels/OPCR/{filename}.xlsx"
    wb.save(link)

    file_url = upload_file(link, "commithub-bucket", f"OPCR/{filename}.xlsx")
    
    return file_url
def createNewIPCR(data, given, middle, last, individuals, position, dates):
    wb = load_workbook("excels/IPCRTest.xlsx")
    ws = wb.active


    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    period = ""
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year
    
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # Set paper size to A4
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Set all margins to 0
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    # Fit all columns in one page (width = 1 page, height = unlimited)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    name = f"{given} {str(middle[0]).capitalize()}. {last}"
    ws["A6"] = f"I, {name}, {position} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = position
    ws["N10"] = formatDate(dates)
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")
    
    ws["A13"] = individuals["review"]["name"]
    ws["A14"] = individuals["review"]["position"]
    ws["H13"] = formatDate(individuals["review"]["date"])
    ws["H13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["H13"].font = Font(bold=True, name="Calibri", size="11")
    
    ws["J13"] = individuals["approve"]["name"]
    ws["J14"] = individuals["approve"]["position"]
    ws["R13"] = formatDate(individuals["approve"]["date"])
    ws["R13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["R13"].font = Font(bold=True, name="Calibri", size="11")
    
    
    row = 26
    starting = "A"
    
    startingrow = 26
    endrow = 0
    numberformat = "0.00"
    

    for i in data:
        if len(i) == 0: continue
            
        for g, h in i.items():
            if len(h) == 0: continue
            print(g) #category name
            ws[str("A"+str(row))] = g #category name
            ws[str("A"+str(row))].font = Font(bold=True, size=12)
            prepareCells(ws, "A"+str(row), str("S"+str(row)))
            row += 1
            
            for a in h:
                print(a)
                #prepare output cell
                print(str("A"+str(row)))
                prepareCells(ws, str("A"+str(row)) ,str("E"+str(row+5)))
                ws[str("A"+str(row))] = a["title"]
                ws[str("A"+str(row))].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                #prepare target cells muna dito 
                prepareCells(ws, str("F"+str(row)), str("F"+str(row + 1)))
                ws["F"+str(row)] = a["summary"]["target"] #quabntity dito
                ws["F"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["F"+str(row)].font = Font(bold=True)
                
                prepareCells(ws, str("F"+str(row+2)), str("F"+str(row + 3)))
                ws["F"+str(row+2)] = a["working_days"]["target"] #days naman dito
                ws["F"+str(row+2)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["F"+str(row+2)].font = Font(bold=True)
                
                prepareCells(ws, str("F"+str(row+4)), str("F"+str(row + 5)))
                ws["F"+str(row+4)] = a["corrections"]["target"] # corrections 
                ws["F"+str(row+4)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["F"+str(row+4)].font = Font(bold=True)
                
                #description
                prepareCells(ws, str("G"+str(row)), str("I"+str(row + 1)))
                ws["G"+str(row)] = a["description"]["target"] # desc quant 
                ws["G"+str(row)].alignment = Alignment(wrap_text=True, horizontal="left", vertical = "center")
                ws.row_dimensions[row].height = 45
                
                prepareCells(ws, str("G"+str(row+2)), str("I"+str(row + 3)))
                ws["G"+str(row+2)] = a["description"]["time"] + " spent" # time quant
                
                prepareCells(ws, str("G"+str(row+4)), str("I"+str(row + 5)))
                ws["G"+str(row+4)] = a["description"]["alterations"]  # correct quant
                
                #prepare dito ng actual cells 
                prepareCells(ws, str("J"+str(row)), str("J"+str(row + 1)))
                ws["J"+str(row)] = a["summary"]["actual"] # desc quant
                ws["J"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["J"+str(row)].font = Font(bold=True)
                
                prepareCells(ws, str("J"+str(row+2)), str("J"+str(row + 3)))
                ws["J"+str(row+2)] = a["working_days"]["actual"] # desc quant
                ws["J"+str(row+2)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["J"+str(row+2)].font = Font(bold=True)
                
                prepareCells(ws, str("J"+str(row+4)), str("J"+str(row + 5)))
                ws["J"+str(row+4)] = a["corrections"]["actual"] # desc quant
                ws["J"+str(row+4)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["J"+str(row+4)].font = Font(bold=True)
                
                prepareCells(ws, str("K"+str(row)), str("M"+str(row + 1)))
                ws["K"+str(row)] = a["description"]["actual"] # desc quant
                ws["K"+str(row)].alignment = Alignment(wrap_text=True, horizontal="left", vertical = "center")
                
                prepareCells(ws, str("K"+str(row+2)), str("M"+str(row + 3)))
                ws["K"+str(row+2)] = a["description"]["time"] + " spent" # desc quant
                
                prepareCells(ws, str("K"+str(row+4)), str("M"+str(row + 5)))
                ws["K"+str(row+4)] = a["description"]["alterations"] # desc quant
                
                #dito yung ratings
                prepareCells(ws, str("N"+str(row)), str("N"+str(row + 5)))
                ws["N"+str(row)] = a["rating"]["quantity"] # quantity
                ws["N"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["N"+str(row)].font = Font(bold=True)
                
                prepareCells(ws, str("O"+str(row)), str("O"+str(row + 5)))
                ws["O"+str(row)] = a["rating"]["efficiency"] # timeliness
                ws["O"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["O"+str(row)].font = Font(bold=True)
                
                prepareCells(ws, str("P"+str(row)), str("P"+str(row + 5)))
                ws["P"+str(row)] = a["rating"]["timeliness"] # corrections
                ws["P"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["P"+str(row)].font = Font(bold=True)
                
                prepareCells(ws, str("Q"+str(row)), str("Q"+str(row + 5)))
                ws["Q"+str(row)] = a["rating"]["average"] # average
                ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws["Q"+str(row)].font = Font(bold=True)
                
                #remarks na wala talagang laman
                prepareCells(ws, str("R"+str(row)), str("S"+str(row + 5)))
                ws["R"+str(row)] = " " # remarks
                
                endrow = row
                row = row + 6
    row += 0
    prepareCells(ws, str("J"+str(row)), str("M"+str(row+1)))
    ws["J"+str(row)] = "Final Average Rating"
    ws["J"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, str("N"+str(row)), str("N"+str(row+1)))
    ws["N"+str(row)] = f"=AVERAGE(N{startingrow+1}: N{endrow})"
    ws["N"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N"+str(row)].number_format = numberformat
    
    prepareCells(ws, str("O"+str(row)), str("O"+str(row+1)))
    ws["O"+str(row)] = f"=AVERAGE(O{startingrow+1}: O{endrow})"
    ws["O"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["O"+str(row)].number_format = numberformat
    
    prepareCells(ws, str("P"+str(row)), str("P"+str(row+1)))
    ws["P"+str(row)] = f"=AVERAGE(P{startingrow+1}: P{endrow})"
    ws["P"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["P"+str(row)].number_format = numberformat
    
    average = f"=AVERAGE(Q{startingrow}: Q{endrow})"
    averagecell = str("Q"+str(row))
    prepareCells(ws, str("Q"+str(row)), str("Q"+str(row+1)))
    ws["Q"+str(row)] = f"=AVERAGE(Q{startingrow+1}: Q{endrow})"
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["Q"+str(row)].number_format = numberformat
    
    #final average rating
    row += 2
    prepareCells(ws, str("J"+str(row)), str("M"+str(row+1)))
    ws["J"+str(row)] = "FINAL AVERAGE RATING"
    ws["J"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, str("N"+str(row)), str("Q"+str(row+1)))
    ws["N"+str(row)] = average
    ws["N"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N"+str(row)].font = Font(bold=True)
    ws["N"+str(row)].number_format = numberformat    
    
    row += 2
    prepareCells(ws, str("J"+str(row)), str("M"+str(row+1)))
    ws["J"+str(row)] = "ADJECTIVAL RATING"
    ws["J"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, str("N"+str(row)), str("Q"+str(row+1)))
    ws["N"+str(row)] = f'=IF(AND({averagecell}>=1, {averagecell}<=1.9), "POOR", IF(AND({averagecell}>=2, {averagecell}<=2.9), "UNSATISFACTORY", IF(AND({averagecell}>=3, {averagecell}<=3.9), "SATISFACTORY", IF(AND({averagecell}>=4, {averagecell}<=4.9), "VERY SATISFACTORY", IF(AND({averagecell}=5), "OUTSTANDING")))))'
    ws["N"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N"+str(row)].font = Font(bold=True)  

    
    #individuals
    row += 3
    prepareCells(ws, str("A"+str(row)), str("D"+str(row)))
    ws["A"+str(row)] = "Discussed With"
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, str("E"+str(row)), str("F"+str(row)))
    ws["E"+str(row)] = "Date"
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, str("G"+str(row)), str("J"+str(row)))
    ws["G"+str(row)] = "Assessed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row)))
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    prepareCells(ws, str("M"+str(row)), str("P"+str(row)))
    ws["M"+str(row)] = "Final Rating By"
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, str("Q"+str(row)), str("S"+str(row)))
    ws["Q"+str(row)] = "Date"
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    
    
    row += 1
    
    #adjust row height for individuals
    # name + position
    prepareCells(ws, str("A"+str(row)), str("D"+str(row+2)))
    ws["A"+str(row)] = admin_data["individuals"]["discuss"]["name"] + "\n" + admin_data["individuals"]["discuss"]["position"]
    ws["A"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["A"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("E"+str(row)), str("F"+str(row+2)))
    ws["E"+str(row)] = ""
    ws["E"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["E"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("G"+str(row)), str("J"+str(row+2)))
    ws["G"+str(row)] = admin_data["individuals"]["assess"]["name"] + "\n" + admin_data["individuals"]["assess"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row+2)))
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("M"+str(row)), str("P"+str(row+2)))
    ws["M"+str(row)] = admin_data["individuals"]["final"]["name"] + "\n" + admin_data["individuals"]["final"]["position"]
    ws["M"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["M"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("Q"+str(row)), str("S"+str(row+2)))
    ws["Q"+str(row)] = ""
    ws["Q"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["Q"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    row+= 4
    
    prepareCells(ws, str("G"+str(row)), str("J"+str(row)))
    ws["G"+str(row)] = "Confirmed By"
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True)
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row)))
    ws["K"+str(row)] = "Date"
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row += 1
    prepareCells(ws, str("G"+str(row)), str("J"+str(row+2)))
    ws["G"+str(row)] = admin_data["individuals"]["confirm"]["name"] + "\n" + admin_data["individuals"]["confirm"]["position"]
    ws["G"+str(row)].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws["G"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    prepareCells(ws, str("K"+str(row)), str("L"+str(row+2)))
    ws["K"+str(row)] = ""
    ws["K"+str(row)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["K"+str(row)].font = Font(bold=True, name="Calibri", size="11")
    
    # Save changes
    id = random.randint(1, 999999)
    filename = f"IPCR-NC-{period}-{given}-{middle}-{last}-{datee}-{id}"
    link = f"excels/IPCR/{filename}.xlsx"
    wb.save(link)
    
    return "downloadlink"
