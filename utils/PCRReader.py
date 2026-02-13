from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
import datetime 

import random
from datetime import date
from app import db

def read_ipcr(file_path):
    from models.Tasks import Sub_Task
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    current_row = 27
    count = 0

    while ws[f'T{current_row}'].value:

        sub_id = int(ws[f'T{current_row}'].value)
        print("SUB TASK ID", sub_id)
        


        

        found_task = Sub_Task.query.get(sub_id)

        if found_task:
            print("CURRENT ROW", current_row)
            count += 1
            quality = float(ws[f'N{current_row}'].value)
            efficiency = float(ws[f'O{current_row}'].value)
            timeliness = float(ws[f'P{current_row}'].value)

            found_task.quantity = quality
            found_task.efficiency = efficiency
            found_task.timeliness = timeliness
            found_task.average = float((quality + efficiency + timeliness) / 3)

            db.session.commit()
        current_row += 6

    if count:
        return "Successfully updated rating data."
    else:
        return "There is no rating data to update."
        

