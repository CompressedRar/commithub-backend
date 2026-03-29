# utils/ipcr_excel.py
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import datetime, random
from utils.FileStorage import upload_file

# Adjust these imports to match your project structure:
from models.PCR import IPCR           # or from models.pcr import IPCR
from models.Tasks import Output, Sub_Task, Main_Task  # adjust path if needed
from models.User import User          # adjust path if needed


def formatDate(d):
    if d is None:
        return ""
    if isinstance(d, str):
        # if stored as string like "YYYY-MM-DD" keep behavior similar to old func
        try:
            parts = str(d).split("-")
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        except Exception:
            return d
    # datetime-like object
    try:
        return d.strftime("%d-%m-%Y")
    except Exception:
        return str(d)


def prepareCells(ws, start, end):
    ws.merge_cells(f"{start}:{end}")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws[f"{start}:{end}"]:
        for cell in row:
            cell.border = border_style


def _build_data_from_ipcr(ipcr, is_draft=False):
    from models.Tasks import Assigned_Department
    from models.System_Settings import System_Settings
    
    settings = System_Settings.get_default_settings()
    
    # Internal grouping dictionary for easy categorization
    grouped = {
        "CORE FUNCTION": {},
        "SUPPORT FUNCTION": {},
        "STRATEGIC FUNCTION": {}
    }

    for output in ipcr.outputs:
        sub = getattr(output, "sub_task", None)
        mt = getattr(output, "main_task", None)
        if not mt or not sub:
            continue

        # 1. Fetch Weight from Department Assignment (Just like OPCR)
        ad = Assigned_Department.query.filter_by(
            main_task_id=mt.id,
            department_id=ipcr.user.department_id,
            period=settings.current_period_id
        ).first()
        weight = (ad.task_weight / 100) if ad else 0

        # 2. Determine Category and Section
        category_name = mt.category.name if mt.category else "Uncategorized"
        section_key = mt.category.type.upper() if mt.category else "CORE FUNCTION"
        
        # Ensure the section exists in our grouped dict (handles casing mismatches)
        if section_key not in grouped:
            grouped[section_key] = {}

        if category_name not in grouped[section_key]:
            grouped[section_key][category_name] = []

        # 3. Draft Mode Logic: Set actuals to 0 if drafting
        actual_acc = 0 if is_draft else getattr(sub, "actual_acc", 0)
        actual_mod = 0 if is_draft else getattr(sub, "actual_mod", 0)
        actual_time = 0 if is_draft else getattr(sub, "actual_time", 0)

        # 4. Rating Calculations
        q = 0 if is_draft else sub.quantity
        e = 0 if is_draft else sub.efficiency
        t = 0 if is_draft else sub.timeliness
        
        avg = (q + e + t) / 3 if (q + e + t) > 0 else 0
        weighted_avg = avg * weight

        # 5. Build Entry "a"
        a = {
            "title": mt.mfo,
            "weight": weight,
            "summary": {
                "target": getattr(sub, "target_acc", 0),
                "actual": actual_acc
            },
            "corrections": {
                "target": getattr(sub, "target_mod", 0),
                "actual": actual_mod
            },
            "sub_task_id": getattr(sub, "id", 0)
        }

        # Timeliness Description Logic
        if mt.timeliness_mode == "timeframe":
            a["working_days"] = {"target": getattr(sub, "target_time", 0), "actual": actual_time}
            target_deadline_desc = "within timeframe"
            actual_deadline_desc = ""
        else:
            # Deadline mode calculation
            days_late = 0
            if not is_draft and sub.actual_deadline and mt.target_deadline:
                days_late = (sub.actual_deadline - mt.target_deadline).days
            
            a["working_days"] = {"target": 1, "actual": abs(days_late)}
            target_deadline_desc = "on the set deadline"
            actual_deadline_desc = "on the set deadline" if days_late == 0 else ("day/s after" if days_late > 0 else "day/s before")

        a["description"] = {
            "target": mt.target_accomplishment,
            "time": mt.time_description,
            "alterations": mt.modification,
            "actual": "" if is_draft else mt.actual_accomplishment,
            "timeliness_mode": mt.timeliness_mode,
            "target_deadline_desc": target_deadline_desc,
            "actual_deadline_desc": actual_deadline_desc if not is_draft else ""
        }

        # 6. Remark Formula Logic
        rating_thresholds = settings.rating_thresholds if settings else {}
        formula_parts = []
        for rating_name, thresholds in rating_thresholds.items():
            min_v, max_v = thresholds.get("min", 0), thresholds.get("max", 5)
            if min_v == max_v:
                formula_parts.append(f'IF({avg}={min_v}, "{str(rating_name).replace("_", " ").upper()}"')
            else:
                formula_parts.append(f'IF(AND({avg}>={min_v}, {avg}<={max_v}), "{str(rating_name).replace("_", " ").upper()}"')
        
        formula = '"UNRATED"' + (")" * len(formula_parts))
        if formula_parts:
            formula = ", ".join(formula_parts) + ", " + formula

        a["rating"] = {
            "quantity": q,
            "efficiency": e,
            "timeliness": t,
            "average": avg,
            "weighted_avg": weighted_avg,
            "remarks": f"={formula}"
        }

        grouped[section_key][category_name].append(a)

    # 7. Final Transformation to maintain the requested structure
    final = {
        "CORE FUNCTION": [],
        "SUPPORT FUNCTION": [],
        "STRATEGIC FUNCTION": []
    }

    for section in grouped:
        for category_name, entries in grouped[section].items():
            final[section].append({category_name: entries})

    return final


def createNewIPCR_from_db(ipcr_id, individuals=None, filename_prefix=None, is_draft = False):

    
    """
    Build an IPCR Excel using DB models. Very similar to your original `createNewIPCR`.
    - ipcr_id: the IPCR DB id to render
    - individuals: dict with same structure your original function expects:
        {
          "review": {"name": "", "position":"", "date": <date>},
          "approve": {...},
          "discuss": {...},
          "assess": {...},
          "final": {...},
          "confirm": {...}
        }
      If None, the function will take placeholders from ipcr fields (if present) or use blank.
    - filename_prefix: optional string to add to filename
    Returns: download link from UploadManager.upload_Report (same pattern as your original).
    """
    wb = load_workbook("excels/IPCRTest.xlsx")
    ws = wb.active

    ipcr = IPCR.query.get(ipcr_id)
    if not ipcr:
        raise ValueError("IPCR not found")

    user = ipcr.user
    if not user:
        raise ValueError("IPCR has no user")

    # If individuals not provided, try to fill from ipcr text fields; otherwise use blanks
    individuals={
        "review": {"name": ipcr.reviewed_by, "position": ipcr.rev_position, "date": ""},
        "approve": {"name": ipcr.approved_by, "position": ipcr.app_position, "date": "" },
        "discuss": {"name": ipcr.discussed_with, "position": ipcr.dis_position, "date": "" },
        "assess": {"name": ipcr.assessed_by, "position": ipcr.ass_position, "date": "" },
        "final": {"name": ipcr.final_rating_by, "position": ipcr.fin_position, "date": "" },
        "confirm": {"name": ipcr.confirmed_by, "position": ipcr.con_position, "date": "" }
    }

    if individuals is None:
        def _mk(name):
            # try to parse ipcr.<name> if available; ipcr fields are plain text in your model
            val = getattr(ipcr, name, "")
            return {"name": val or "", "position": "", "date": None}
        individuals = {
            "review": _mk("reviewed_by"),
            "approve": _mk("approved_by"),
            "discuss": _mk("discussed_with"),
            "assess": _mk("assessed_by"),
            "final": _mk("final_rating_by"),
            "confirm": _mk("confirmed_by")
        }

    print(individuals)

    

    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year

    # page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = ws.page_margins.right = 0
    ws.page_margins.top = ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # name formatting (same as your old function)
    given = user.first_name
    middle = user.middle_name or ""
    last = user.last_name
    mid_initial = str(middle[0]).capitalize() if middle else ""
    name = f"{given} {mid_initial}. {last}"
    position = ""  # you stored user.position relation; adjust if needed


    try:
        position = user.position.info()["name"] if user.position else ""
    except Exception:
        # fallback to empty or text field
        position = getattr(user, "position", "") or ""

    # header fields
    ws["A6"] = f"I, {name}, {position} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = position
    ws["N10"] = ""
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")
    
    
    #metadata here
    ws["A8"] = ipcr_id
    print("EXCEL IPCR ID", ipcr_id)
    ws["A8"].font = Font(bold=True, name="Calibri", size="1", color="FFFFFF")

    ws["A9"] = user.id
    print("USERID", user.id)
    ws["A9"].font = Font(bold=True, name="Calibri", size="1", color="FFFFFF")
    

    # individuals: review / approve
    ws["A13"] = individuals.get("review", {}).get("name", "")
    ws["A14"] = individuals.get("review", {}).get("position", "")
    ws["H13"] = ""
    ws["H13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["H13"].font = Font(bold=True, name="Calibri", size="11")

    ws["J13"] = individuals.get("approve", {}).get("name", "")
    ws["J14"] = individuals.get("approve", {}).get("position", "")
    ws["R13"] = ""
    ws["R13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["R13"].font = Font(bold=True, name="Calibri", size="11")

    # main loop writing tasks (kept identical to your original)
    row = 26
    startingrow = 26
    endrow = 0
    numberformat = "0.00"

    ipcr_data = _build_data_from_ipcr(ipcr)

    for type, data in ipcr_data.items():
        
        print("TYPE", type)
        total = 0
        task_frequency = 0
        if type != "CORE FUNCTION" and len(data) != 0:         

            ws[f"A{row}"] = type
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        

        for i in data:
            if len(i) == 0:
                continue
            

            for g, h in i.items():
                if len(h) == 0:
                    continue

                ws[f"A{row}"] = g  # category name
                prepareCells(ws, f"A{row}", f"S{row}")
                row += 1

                for a in h:
                    task_frequency += 1
                    print("TASKS", a)
                    prepareCells(ws, f"A{row}", f"E{row+5}")
                    ws[f"A{row}"] = a["title"]
                    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # targets
                    prepareCells(ws, f"F{row}", f"F{row+1}")
                    ws[f"F{row}"] = a["summary"]["target"]
                    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"F{row}"].font = Font(bold=True)

                    # target deadline / timeframe

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"F{row+2}", f"F{row+3}")
                        ws[f"F{row+2}"] = f""
                        ws[f"F{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"F{row+2}"].font = Font(bold=True)

                    else:
                        prepareCells(ws, f"F{row+2}", f"F{row+3}")
                        ws[f"F{row+2}"] = a["working_days"]["target"]
                        ws[f"F{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"F{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"F{row+4}", f"F{row+5}")
                    ws[f"F{row+4}"] = a["corrections"]["target"]
                    ws[f"F{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"F{row+4}"].font = Font(bold=True)

                    # description target
                    prepareCells(ws, f"G{row}", f"I{row+1}")
                    ws[f"G{row}"] = a["description"]["target"]
                    ws[f"G{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws.row_dimensions[row].height = 45

                    # target deadline / timeframe
                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"G{row+2}", f"I{row+3}")
                        ws[f"G{row+2}"] = f"{a['description']['target_deadline_desc']}"
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    else:
                        prepareCells(ws, f"G{row+2}", f"I{row+3}")
                        ws[f"G{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    prepareCells(ws, f"G{row+4}", f"I{row+5}")
                    ws[f"G{row+4}"] = a["description"]["alterations"]
                    ws[f"G{row+4}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # actual cells
                    prepareCells(ws, f"J{row}", f"J{row+1}")
                    ws[f"J{row}"] = 0 if is_draft else a["summary"]["actual"]
                    ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row}"].font = Font(bold=True)

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = 0 if is_draft else a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)
                    else:
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = 0 if is_draft else a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"J{row+4}", f"J{row+5}")
                    ws[f"J{row+4}"] = 0 if is_draft else a["corrections"]["actual"]
                    ws[f"J{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+4}"].font = Font(bold=True)

                    prepareCells(ws, f"K{row}", f"M{row+1}")
                    ws[f"K{row}"] = a["description"]["actual"]
                    ws[f"K{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = f"{a['description']['actual_deadline_desc']}"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    else:
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    prepareCells(ws, f"K{row+4}", f"M{row+5}")
                    ws[f"K{row+4}"] = a["description"]["alterations"]
                    ws[f"K{row+4}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    # ratings
                    prepareCells(ws, f"N{row}", f"N{row+5}")
                    ws[f"N{row}"] = 0 if is_draft else a["rating"]["quantity"]
                    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"N{row}"].font = Font(bold=True)

                    prepareCells(ws, f"O{row}", f"O{row+5}")
                    ws[f"O{row}"] = 0 if is_draft else a["rating"]["efficiency"]
                    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"O{row}"].font = Font(bold=True)

                    prepareCells(ws, f"P{row}", f"P{row+5}")
                    ws[f"P{row}"] = 0 if is_draft else a["rating"]["timeliness"]
                    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"P{row}"].font = Font(bold=True)

                    total += a["rating"]["average"]

                    prepareCells(ws, f"Q{row}", f"Q{row+5}")
                    ws[f"Q{row}"] = 0 if is_draft else f"=AVERAGE(N{row}, O{row},P{row})"
                    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"Q{row}"].font = Font(bold=True)
                    ws[f"Q{row}"].number_format = numberformat

                    # remarks (blank)
                    prepareCells(ws, f"R{row}", f"S{row+5}")
                    ws[f"R{row}"] = ""
                    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"R{row}"].font = Font(bold=True)
                    
                    ws[f"T{row}"] = a["sub_task_id"]
                    ws[f"T{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"T{row}"].font = Font(bold=True, color = "ffffff", size=1)
                    

                    endrow = row
                    row = row + 6        
            
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "Final Average Rating"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"N{row+1}")
    ws[f"N{row}"] = 0 if is_draft else f"=AVERAGE(N{startingrow}: N{endrow})"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].number_format = numberformat

    prepareCells(ws, f"O{row}", f"O{row+1}")
    ws[f"O{row}"] = 0 if is_draft else f"=AVERAGE(O{startingrow}: O{endrow})"
    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"O{row}"].number_format = numberformat

    prepareCells(ws, f"P{row}", f"P{row+1}")
    ws[f"P{row}"] = 0 if is_draft else f"=AVERAGE(P{startingrow}: P{endrow})"
    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"P{row}"].number_format = numberformat

    average = f"=AVERAGE(Q{startingrow}: Q{endrow})"
    averagecell = f"Q{row}"
    prepareCells(ws, f"Q{row}", f"Q{row+1}")
    ws[f"Q{row}"] = 0 if is_draft else f"=AVERAGE(Q{startingrow}: Q{endrow})"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].number_format = numberformat

    # final average block
    row += 2
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "FINAL AVERAGE RATING"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = 0 if is_draft else average
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)
    ws[f"N{row}"].number_format = numberformat

    row += 2
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "ADJECTIVAL RATING"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    from models.System_Settings import System_Settings
    settings = System_Settings.get_default_settings()
    rating_thresholds = settings.rating_thresholds if settings else {
        "POOR": {"min": 0, "max": 1.49},
        "UNSATISFACTORY": {"min": 1.5, "max": 2.49},
        "SATISFACTORY": {"min": 2.5, "max": 3.49},
        "VERY SATISFACTORY": {"min": 3.5, "max": 4.49},
        "OUTSTANDING": {"min": 4.5, "max": 5}
    }

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    
    # Build dynamic formula from rating thresholds
    formula_parts = []
    for rating_name, thresholds in rating_thresholds.items():
        min_val = thresholds.get("min", 0)
        max_val = thresholds.get("max", 5)
        if min_val == max_val:
            # Exact match (e.g., OUTSTANDING = 5)
            formula_parts.append(f'IF({averagecell}={min_val}, "{str(rating_name).replace("_", " ").upper()}"')
        else:
            # Range match
            formula_parts.append(f'IF(AND({averagecell}>={min_val}, {averagecell}<={max_val}), "{str(rating_name).replace("_", " ").upper()}"')
    
    # Close all nested IFs
    formula = ""
    for i, part in enumerate(formula_parts):
        formula += part + ", "
    formula += '"UNRATED"' + ")" * len(formula_parts)
    
    ws[f"N{row}"] = f"={formula}"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)

    # individuals block (Discussed, Assessed, Final, Confirm)
    row += 3
    prepareCells(ws, f"A{row}", f"D{row}")
    ws[f"A{row}"] = "Discussed With"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True)

    prepareCells(ws, f"E{row}", f"F{row}")
    ws[f"E{row}"] = "Date"
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Assessed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"M{row}", f"P{row}")
    ws[f"M{row}"] = "Final Rating By"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True)

    prepareCells(ws, f"Q{row}", f"S{row}")
    ws[f"Q{row}"] = "Date"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    # name + position for discussed/assess/final
    prepareCells(ws, f"A{row}", f"D{row+2}")
    ws[f"A{row}"] = f"{individuals.get('discuss', {}).get('name','')}\n{individuals.get('discuss', {}).get('position','')}"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"E{row}", f"F{row+2}")
    ws[f"E{row}"] = ""
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"E{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('assess', {}).get('name','')}\n{individuals.get('assess', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"M{row}", f"P{row+2}")
    ws[f"M{row}"] = f"{individuals.get('final', {}).get('name','')}\n{individuals.get('final', {}).get('position','')}"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"Q{row}", f"S{row+2}")
    ws[f"Q{row}"] = ""
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].font = Font(bold=True, name="Calibri", size="11")

    row += 4
    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Confirmed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('confirm', {}).get('name','')}\n{individuals.get('confirm', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    # Save file
    id_rand = random.randint(1, 999999)
    fname_given = given
    fname_middle = middle[0] if middle else ""
    fname_last = last
    prefix = filename_prefix if filename_prefix else "IPCR"
    filename = f"{prefix} {period} - {fname_last} - {id_rand}"
    link = f"excels/IPCR/{filename}.xlsx"
    wb.save(link)

    file_url = upload_file(link, "commithub-bucket", f"IPCR/{filename}.xlsx")


    return file_url


def createNewWeightedIPCR_from_db(ipcr_id, individuals=None, filename_prefix=None):

    
    """
    Build an IPCR Excel using DB models. Very similar to your original `createNewIPCR`.
    - ipcr_id: the IPCR DB id to render
    - individuals: dict with same structure your original function expects:
        {
          "review": {"name": "", "position":"", "date": <date>},
          "approve": {...},
          "discuss": {...},
          "assess": {...},
          "final": {...},
          "confirm": {...}
        }
      If None, the function will take placeholders from ipcr fields (if present) or use blank.
    - filename_prefix: optional string to add to filename
    Returns: download link from UploadManager.upload_Report (same pattern as your original).
    """

    ipcr = IPCR.query.get(ipcr_id)
    if not ipcr:
        raise ValueError("IPCR not found")

    user = ipcr.user
    if not user:
        raise ValueError("IPCR has no user")

    # If individuals not provided, try to fill from ipcr text fields; otherwise use blanks
    individuals={
        "review": {"name": ipcr.reviewed_by, "position": ipcr.rev_position, "date": ""},
        "approve": {"name": ipcr.approved_by, "position": ipcr.app_position, "date": "" },
        "discuss": {"name": ipcr.discussed_with, "position": ipcr.dis_position, "date": "" },
        "assess": {"name": ipcr.assessed_by, "position": ipcr.ass_position, "date": "" },
        "final": {"name": ipcr.final_rating_by, "position": ipcr.fin_position, "date": "" },
        "confirm": {"name": ipcr.confirmed_by, "position": ipcr.con_position, "date": "" }
    }

    if individuals is None:
        def _mk(name):
            # try to parse ipcr.<name> if available; ipcr fields are plain text in your model
            val = getattr(ipcr, name, "")
            return {"name": val or "", "position": "", "date": None}
        individuals = {
            "review": _mk("reviewed_by"),
            "approve": _mk("approved_by"),
            "discuss": _mk("discussed_with"),
            "assess": _mk("assessed_by"),
            "final": _mk("final_rating_by"),
            "confirm": _mk("confirmed_by")
        }

    # Build 'data' structure from DB outputs/subtasks
    

    # Now replicate your excel logic exactly (with the same cell positions & merges)
    wb = load_workbook("excels/IPCRTest.xlsx")
    ws = wb.active

    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year

    # page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = ws.page_margins.right = 0
    ws.page_margins.top = ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # name formatting (same as your old function)
    given = user.first_name
    middle = user.middle_name or ""
    last = user.last_name
    mid_initial = str(middle[0]).capitalize() if middle else ""
    name = f"{given} {mid_initial}. {last}"
    position = ""  # you stored user.position relation; adjust if needed

    position_weights = user.position.info() if user.position else {
        "core_weight": 1.0,
        "strategic_weight": 1.0,
        "support_weight": 1.0
    }

    try:
        position = user.position.info()["name"] if user.position else ""
    except Exception:
        # fallback to empty or text field
        position = getattr(user, "position", "") or ""

    # header fields
    ws["A6"] = f"I, {name}, {position} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = position
    ws["N10"] = ""
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")

    # individuals: review / approve
    ws["A13"] = individuals.get("review", {}).get("name", "")
    ws["A14"] = individuals.get("review", {}).get("position", "")
    ws["H13"] = ""
    ws["H13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["H13"].font = Font(bold=True, name="Calibri", size="11")

    ws["J13"] = individuals.get("approve", {}).get("name", "")
    ws["J14"] = individuals.get("approve", {}).get("position", "")
    ws["R13"] = ""
    ws["R13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["R13"].font = Font(bold=True, name="Calibri", size="11")

    # main loop writing tasks (kept identical to your original)
    row = 26
    startingrow = 26
    endrow = 0
    numberformat = "0.00"

    ipcr_data = _build_data_from_ipcr(ipcr)

    for type, data in ipcr_data.items():
        
        print("TYPE", type)
        total = 0
        if type != "CORE FUNCTION" and len(data) != 0:         

            ws[f"A{row}"] = type
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        

        for i in data:
            if len(i) == 0:
                continue
            

            for g, h in i.items():
                if len(h) == 0:
                    continue

                ws[f"A{row}"] = g  # category name
                prepareCells(ws, f"A{row}", f"S{row}")
                row += 1

                for a in h:
                    print("TASKS", a)
                    prepareCells(ws, f"A{row}", f"E{row+5}")
                    ws[f"A{row}"] = a["title"]
                    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # targets
                    prepareCells(ws, f"F{row}", f"F{row+1}")
                    ws[f"F{row}"] = a["summary"]["target"]
                    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"F{row}"].font = Font(bold=True)

                    # target deadline / timeframe

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"F{row+2}", f"F{row+3}")
                        ws[f"F{row+2}"] = f""
                        ws[f"F{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"F{row+2}"].font = Font(bold=True)

                    else:
                        prepareCells(ws, f"F{row+2}", f"F{row+3}")
                        ws[f"F{row+2}"] = a["working_days"]["target"]
                        ws[f"F{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"F{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"F{row+4}", f"F{row+5}")
                    ws[f"F{row+4}"] = a["corrections"]["target"]
                    ws[f"F{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"F{row+4}"].font = Font(bold=True)

                    # description target
                    prepareCells(ws, f"G{row}", f"I{row+1}")
                    ws[f"G{row}"] = a["description"]["target"]
                    ws[f"G{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws.row_dimensions[row].height = 45

                    # target deadline / timeframe
                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"G{row+2}", f"I{row+3}")
                        ws[f"G{row+2}"] = f"{a['description']['target_deadline_desc']}"
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    else:
                        prepareCells(ws, f"G{row+2}", f"I{row+3}")
                        ws[f"G{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    prepareCells(ws, f"G{row+4}", f"I{row+5}")
                    ws[f"G{row+4}"] = a["description"]["alterations"]
                    ws[f"G{row+4}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # actual cells
                    prepareCells(ws, f"J{row}", f"J{row+1}")
                    ws[f"J{row}"] = a["summary"]["actual"]
                    ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row}"].font = Font(bold=True)

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)
                    else:
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"J{row+4}", f"J{row+5}")
                    ws[f"J{row+4}"] = a["corrections"]["actual"]
                    ws[f"J{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+4}"].font = Font(bold=True)

                    prepareCells(ws, f"K{row}", f"M{row+1}")
                    ws[f"K{row}"] = a["description"]["actual"]
                    ws[f"K{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = f"{a['description']['actual_deadline_desc']}"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    else:
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    prepareCells(ws, f"K{row+4}", f"M{row+5}")
                    ws[f"K{row+4}"] = a["description"]["alterations"]
                    ws[f"K{row+4}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    # ratings
                    prepareCells(ws, f"N{row}", f"N{row+5}")
                    ws[f"N{row}"] = a["rating"]["quantity"]
                    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"N{row}"].font = Font(bold=True)

                    prepareCells(ws, f"O{row}", f"O{row+5}")
                    ws[f"O{row}"] = a["rating"]["efficiency"]
                    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"O{row}"].font = Font(bold=True)

                    prepareCells(ws, f"P{row}", f"P{row+5}")
                    ws[f"P{row}"] = a["rating"]["timeliness"]
                    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"P{row}"].font = Font(bold=True)

                    prepareCells(ws, f"Q{row}", f"Q{row+5}")
                    ws[f"Q{row}"] = a["rating"]["average"]
                    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"Q{row}"].font = Font(bold=True)

                    total += a["rating"]["average"]

                    # remarks (blank)
                    prepareCells(ws, f"R{row}", f"S{row+5}")
                    ws[f"R{row}"] =a["rating"]["remarks"]
                    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"R{row}"].font = Font(bold=True)

                    endrow = row
                    row = row + 6

        raw_average = total / len(data) if len(data) > 0 else 0
        
        ws[f"A{row}"] = "RAW AVERAGE: " + f"{raw_average}" if len(data) > 0 else "RAW AVERAGE: 0"
        prepareCells(ws, f"A{row}", f"S{row}")
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        if type == "CORE FUNCTION":
            core_weight = position_weights.get("core_weight", 1.0)
            weighted_average = raw_average * core_weight
            total = weighted_average * len(data)  # for final average calculation

            ws[f"A{row}"] = "WEIGHTED AVERAGE: " + f"{weighted_average:.0f}" if len(data) > 0 else "RAW AVERAGE: 0"
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        elif type == "STRATEGIC FUNCTION":
            strategic_weight = position_weights.get("strategic_weight", 1.0)
            weighted_average = raw_average * strategic_weight
            total = weighted_average * len(data)  # for final average calculation

            ws[f"A{row}"] = "WEIGHTED AVERAGE: " + f"{weighted_average:.0f}" if len(data) > 0 else "RAW AVERAGE: 0"
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        elif type == "SUPPORT FUNCTION":
            support_weight = position_weights.get("support_weight", 1.0)
            weighted_average = raw_average * support_weight
            total = weighted_average * len(data)  # for final average calculation

            ws[f"A{row}"] = "WEIGHTED AVERAGE: " + f"{weighted_average:.0f}" if len(data) > 0 else "RAW AVERAGE: 0"
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1


    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "Final Average Rating"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"N{row+1}")
    ws[f"N{row}"] = f"=AVERAGE(N{startingrow}: N{endrow})"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].number_format = numberformat

    prepareCells(ws, f"O{row}", f"O{row+1}")
    ws[f"O{row}"] = f"=AVERAGE(O{startingrow}: O{endrow})"
    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"O{row}"].number_format = numberformat

    prepareCells(ws, f"P{row}", f"P{row+1}")
    ws[f"P{row}"] = f"=AVERAGE(P{startingrow}: P{endrow})"
    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"P{row}"].number_format = numberformat

    average = f"=AVERAGE(Q{startingrow}: Q{endrow})"
    averagecell = f"Q{row}"
    prepareCells(ws, f"Q{row}", f"Q{row+1}")
    ws[f"Q{row}"] = f"=AVERAGE(Q{startingrow}: Q{endrow})"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].number_format = numberformat

    # final average block
    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "FINAL AVERAGE RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = average
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)
    ws[f"N{row}"].number_format = numberformat

    row += 2
    prepareCells(ws, f"L{row}", f"M{row+1}")
    ws[f"L{row}"] = "ADJECTIVAL RATING"
    ws[f"L{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    from models.System_Settings import System_Settings
    settings = System_Settings.get_default_settings()
    rating_thresholds = settings.rating_thresholds if settings else {
        "POOR": {"min": 1, "max": 1.9},
        "UNSATISFACTORY": {"min": 2, "max": 2.9},
        "SATISFACTORY": {"min": 3, "max": 3.9},
        "VERY SATISFACTORY": {"min": 4, "max": 4.9},
        "OUTSTANDING": {"min": 5, "max": 5}
    }

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    
    # Build dynamic formula from rating thresholds
    formula_parts = []
    for rating_name, thresholds in rating_thresholds.items():
        min_val = thresholds.get("min", 0)
        max_val = thresholds.get("max", 5)
        if min_val == max_val:
            # Exact match (e.g., OUTSTANDING = 5)
            formula_parts.append(f'IF({averagecell}={min_val}, "{str(rating_name).replace("_", " ").upper()}"')
        else:
            # Range match
            formula_parts.append(f'IF(AND({averagecell}>={min_val}, {averagecell}<={max_val}), "{str(rating_name).replace("_", " ").upper()}"')
    
    # Close all nested IFs
    formula = ""
    for i, part in enumerate(formula_parts):
        formula += part + ", "
    formula += '"UNRATED"' + ")" * len(formula_parts)
    
    ws[f"N{row}"] = f"={formula}"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)

    # individuals block (Discussed, Assessed, Final, Confirm)
    row += 3
    prepareCells(ws, f"A{row}", f"D{row}")
    ws[f"A{row}"] = "Discussed With"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True)

    prepareCells(ws, f"E{row}", f"F{row}")
    ws[f"E{row}"] = "Date"
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Assessed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"M{row}", f"P{row}")
    ws[f"M{row}"] = "Final Rating By"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True)

    prepareCells(ws, f"Q{row}", f"S{row}")
    ws[f"Q{row}"] = "Date"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    # name + position for discussed/assess/final
    prepareCells(ws, f"A{row}", f"D{row+2}")
    ws[f"A{row}"] = f"{individuals.get('discuss', {}).get('name','')}\n{individuals.get('discuss', {}).get('position','')}"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"E{row}", f"F{row+2}")
    ws[f"E{row}"] = ""
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"E{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('assess', {}).get('name','')}\n{individuals.get('assess', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"M{row}", f"P{row+2}")
    ws[f"M{row}"] = f"{individuals.get('final', {}).get('name','')}\n{individuals.get('final', {}).get('position','')}"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"Q{row}", f"S{row+2}")
    ws[f"Q{row}"] = ""
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].font = Font(bold=True, name="Calibri", size="11")

    row += 4
    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Confirmed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('confirm', {}).get('name','')}\n{individuals.get('confirm', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    # Save file
    id_rand = random.randint(1, 999999)
    fname_given = given
    fname_middle = middle[0] if middle else ""
    fname_last = last
    prefix = filename_prefix if filename_prefix else "IPCR"
    filename = f"{prefix} {period} - {fname_last} - {id_rand}"
    link = f"excels/IPCR/{filename}.xlsx"
    wb.save(link)

    file_url = upload_file(link, "commithub-bucket", f"IPCR/{filename}.xlsx")


    return file_url


def createWeightedIPCR_from_db(ipcr_id, individuals=None, filename_prefix=None, is_draft = False):

    
    """
    Build an IPCR Excel using DB models. Very similar to your original `createNewIPCR`.
    - ipcr_id: the IPCR DB id to render
    - individuals: dict with same structure your original function expects:
        {
          "review": {"name": "", "position":"", "date": <date>},
          "approve": {...},
          "discuss": {...},
          "assess": {...},
          "final": {...},
          "confirm": {...}
        }
      If None, the function will take placeholders from ipcr fields (if present) or use blank.
    - filename_prefix: optional string to add to filename
    Returns: download link from UploadManager.upload_Report (same pattern as your original).
    """
    wb = load_workbook("excels/WeightedIPCRTest.xlsx")
    ws = wb.active

    ipcr = IPCR.query.get(ipcr_id)
    if not ipcr:
        raise ValueError("IPCR not found")

    user = ipcr.user
    if not user:
        raise ValueError("IPCR has no user")

    # If individuals not provided, try to fill from ipcr text fields; otherwise use blanks
    individuals={
        "review": {"name": ipcr.reviewed_by, "position": ipcr.rev_position, "date": ""},
        "approve": {"name": ipcr.approved_by, "position": ipcr.app_position, "date": "" },
        "discuss": {"name": ipcr.discussed_with, "position": ipcr.dis_position, "date": "" },
        "assess": {"name": ipcr.assessed_by, "position": ipcr.ass_position, "date": "" },
        "final": {"name": ipcr.final_rating_by, "position": ipcr.fin_position, "date": "" },
        "confirm": {"name": ipcr.confirmed_by, "position": ipcr.con_position, "date": "" }
    }

    if individuals is None:
        def _mk(name):
            # try to parse ipcr.<name> if available; ipcr fields are plain text in your model
            val = getattr(ipcr, name, "")
            return {"name": val or "", "position": "", "date": None}
        individuals = {
            "review": _mk("reviewed_by"),
            "approve": _mk("approved_by"),
            "discuss": _mk("discussed_with"),
            "assess": _mk("assessed_by"),
            "final": _mk("final_rating_by"),
            "confirm": _mk("confirmed_by")
        }

    print(individuals)

    

    datee = datetime.datetime.now().month
    year = str(datetime.datetime.now().year)
    if datee <= 6:
        period = "JANUARY - JUNE " + year
    else:
        period = "JULY - DECEMBER " + year

    # page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = ws.page_margins.right = 0
    ws.page_margins.top = ws.page_margins.bottom = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # name formatting (same as your old function)
    given = user.first_name
    middle = user.middle_name or ""
    last = user.last_name
    mid_initial = str(middle[0]).capitalize() if middle else ""
    name = f"{given} {mid_initial}. {last}"
    position = ""  # you stored user.position relation; adjust if needed


    try:
        position = user.position.info()["name"] if user.position else ""
    except Exception:
        # fallback to empty or text field
        position = getattr(user, "position", "") or ""

    # header fields
    ws["A6"] = f"I, {name}, {position} of the  NORZAGARAY COLLEGE, commit to deliver and agree to be rated on the attainment of "
    ws["A7"] = f"the following targets in accordance with the indicated measures for the period of {period}."
    ws["N8"] = name
    ws["N9"] = position
    ws["N10"] = ""
    ws["N10"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["N10"].font = Font(bold=True, name="Calibri", size="11")
    
    
    #metadata here
    ws["A8"] = ipcr_id
    print("EXCEL IPCR ID", ipcr_id)
    ws["A8"].font = Font(bold=True, name="Calibri", size="1", color="FFFFFF")

    ws["A9"] = user.id
    print("USERID", user.id)
    ws["A9"].font = Font(bold=True, name="Calibri", size="1", color="FFFFFF")
    

    # individuals: review / approve
    ws["A13"] = individuals.get("review", {}).get("name", "")
    ws["A14"] = individuals.get("review", {}).get("position", "")
    ws["H13"] = ""
    ws["H13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["H13"].font = Font(bold=True, name="Calibri", size="11")

    ws["J13"] = individuals.get("approve", {}).get("name", "")
    ws["J14"] = individuals.get("approve", {}).get("position", "")
    ws["R13"] = ""
    ws["R13"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["R13"].font = Font(bold=True, name="Calibri", size="11")

    # main loop writing tasks (kept identical to your original)
    row = 26
    startingrow = 26
    endrow = 0
    numberformat = "0.00"

    ipcr_data = _build_data_from_ipcr(ipcr)

    for type, data in ipcr_data.items():
        
        print("TYPE", type)
        total = 0
        task_frequency = 0
        if type != "CORE FUNCTION" and len(data) != 0:         

            ws[f"A{row}"] = type
            prepareCells(ws, f"A{row}", f"S{row}")
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
        

        for i in data:
            if len(i) == 0:
                continue
            

            for g, h in i.items():
                if len(h) == 0:
                    continue

                ws[f"A{row}"] = g  # category name
                prepareCells(ws, f"A{row}", f"S{row}")
                row += 1

                for a in h:
                    task_frequency += 1
                    print("TASKS", a)
                    prepareCells(ws, f"A{row}", f"E{row+5}")
                    ws[f"A{row}"] = a["title"]
                    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # targets

                    prepareCells(ws, f"F{row}", f"F{row+5}")
                    ws[f"F{row}"] = a["weight"]
                    ws[f"F{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"F{row}"].font = Font(bold=True)

                    prepareCells(ws, f"G{row}", f"G{row+1}")
                    ws[f"G{row}"] = a["summary"]["target"]
                    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"G{row}"].font = Font(bold=True)

                    # target deadline / timeframe

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"G{row+2}", f"G{row+3}")
                        ws[f"G{row+2}"] = f""
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"G{row+2}"].font = Font(bold=True)

                    else:
                        prepareCells(ws, f"G{row+2}", f"G{row+3}")
                        ws[f"G{row+2}"] = a["working_days"]["target"]
                        ws[f"G{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"G{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"G{row+4}", f"G{row+5}")
                    ws[f"G{row+4}"] = a["corrections"]["target"]
                    ws[f"G{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"G{row+4}"].font = Font(bold=True)

                    # description target
                    prepareCells(ws, f"H{row}", f"I{row+1}")
                    ws[f"H{row}"] = a["description"]["target"]
                    ws[f"H{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    ws.row_dimensions[row].height = 45

                    # target deadline / timeframe
                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"H{row+2}", f"I{row+3}")
                        ws[f"H{row+2}"] = f"{a['description']['target_deadline_desc']}"
                        ws[f"H{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    else:
                        prepareCells(ws, f"H{row+2}", f"I{row+3}")
                        ws[f"H{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"H{row+2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    prepareCells(ws, f"H{row+4}", f"I{row+5}")
                    ws[f"H{row+4}"] = a["description"]["alterations"]
                    ws[f"H{row+4}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                    # actual cells
                    prepareCells(ws, f"J{row}", f"J{row+1}")
                    ws[f"J{row}"] = 0 if is_draft else a["summary"]["actual"]
                    ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row}"].font = Font(bold=True)

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = 0 if is_draft else a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)
                    else:
                        prepareCells(ws, f"J{row+2}", f"J{row+3}")
                        ws[f"J{row+2}"] = 0 if is_draft else a["working_days"]["actual"]
                        ws[f"J{row+2}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        ws[f"J{row+2}"].font = Font(bold=True)

                    prepareCells(ws, f"J{row+4}", f"J{row+5}")
                    ws[f"J{row+4}"] = 0 if is_draft else a["corrections"]["actual"]
                    ws[f"J{row+4}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"J{row+4}"].font = Font(bold=True)

                    prepareCells(ws, f"K{row}", f"M{row+1}")
                    ws[f"K{row}"] = a["description"]["actual"]
                    ws[f"K{row}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    if a["description"]["timeliness_mode"] == "deadline":
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = f"{a['description']['actual_deadline_desc']}"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    else:
                        prepareCells(ws, f"K{row+2}", f"M{row+3}")
                        ws[f"K{row+2}"] = str(a["description"]["time"] or "") + " with"
                        ws[f"K{row+2}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    prepareCells(ws, f"K{row+4}", f"M{row+5}")
                    ws[f"K{row+4}"] = a["description"]["alterations"]
                    ws[f"K{row+4}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

                    # ratings
                    prepareCells(ws, f"N{row}", f"N{row+5}")
                    ws[f"N{row}"] = 0 if is_draft else a["rating"]["quantity"]
                    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"N{row}"].font = Font(bold=True)

                    prepareCells(ws, f"O{row}", f"O{row+5}")
                    ws[f"O{row}"] = 0 if is_draft else a["rating"]["efficiency"]
                    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"O{row}"].font = Font(bold=True)

                    prepareCells(ws, f"P{row}", f"P{row+5}")
                    ws[f"P{row}"] = 0 if is_draft else a["rating"]["timeliness"]
                    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"P{row}"].font = Font(bold=True)

                    total += a["rating"]["average"]

                    prepareCells(ws, f"Q{row}", f"Q{row+5}")
                    ws[f"Q{row}"] = 0 if is_draft else f"=AVERAGE(N{row}, O{row},P{row})"
                    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"Q{row}"].font = Font(bold=True)
                    ws[f"Q{row}"].number_format = numberformat

                    # remarks (blank)
                    prepareCells(ws, f"R{row}", f"R{row+5}")
                    ws[f"R{row}"] = 0 if is_draft else float(f"{a["rating"]["weighted_avg"]:.2F}")
                    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"R{row}"].font = Font(bold=True)

                    prepareCells(ws, f"S{row}", f"S{row+5}")
                    ws[f"S{row}"] = ""
                    ws[f"S{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"S{row}"].font = Font(bold=True)
                    
                    ws[f"T{row}"] = a["sub_task_id"]
                    ws[f"T{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws[f"T{row}"].font = Font(bold=True, color = "ffffff", size=1)
                    

                    endrow = row
                    row = row + 6        
            
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "Final Average Rating"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"N{row+1}")
    ws[f"N{row}"] = 0 if is_draft else f"=AVERAGE(N{startingrow}: N{endrow})"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].number_format = numberformat

    prepareCells(ws, f"O{row}", f"O{row+1}")
    ws[f"O{row}"] = 0 if is_draft else f"=AVERAGE(O{startingrow}: O{endrow})"
    ws[f"O{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"O{row}"].number_format = numberformat

    prepareCells(ws, f"P{row}", f"P{row+1}")
    ws[f"P{row}"] = 0 if is_draft else f"=AVERAGE(P{startingrow}: P{endrow})"
    ws[f"P{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"P{row}"].number_format = numberformat

    average = f"=AVERAGE(Q{startingrow}: Q{endrow})"
    averagecell = f"Q{row}"
    prepareCells(ws, f"Q{row}", f"Q{row+1}")
    ws[f"Q{row}"] = 0 if is_draft else f"=AVERAGE(Q{startingrow}: Q{endrow})"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].number_format = numberformat

    prepareCells(ws, f"R{row}", f"R{row+1}")
    ws[f"R{row}"] = 0 if is_draft else f"=SUM(R{startingrow+1}: R{endrow})"
    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"R{row}"].number_format = numberformat

    # final average block
    row += 2
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "FINAL AVERAGE RATING"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    ws[f"N{row}"] = 0 if is_draft else average
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)
    ws[f"N{row}"].number_format = numberformat

    prepareCells(ws, f"R{row}", f"R{row+1}")
    ws[f"R{row}"] = 0 if is_draft else f"=SUM(R{startingrow+1}: R{endrow})"
    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"R{row}"].number_format = numberformat

    row += 2
    prepareCells(ws, f"K{row}", f"M{row+1}")
    ws[f"K{row}"] = "ADJECTIVAL RATING"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    from models.System_Settings import System_Settings
    settings = System_Settings.get_default_settings()
    rating_thresholds = settings.rating_thresholds if settings else {
        "POOR": {"min": 0, "max": 1.49},
        "UNSATISFACTORY": {"min": 1.5, "max": 2.49},
        "SATISFACTORY": {"min": 2.5, "max": 3.49},
        "VERY SATISFACTORY": {"min": 3.5, "max": 4.49},
        "OUTSTANDING": {"min": 4.5, "max": 5}
    }

    prepareCells(ws, f"N{row}", f"Q{row+1}")
    prepareCells(ws, f"R{row}", f"R{row+1}")
    
    # Build dynamic formula from rating thresholds
    formula_parts = []
    for rating_name, thresholds in rating_thresholds.items():
        min_val = thresholds.get("min", 0)
        max_val = thresholds.get("max", 5)
        if min_val == max_val:
            # Exact match (e.g., OUTSTANDING = 5)
            formula_parts.append(f'IF({averagecell}={min_val}, "{str(rating_name).replace("_", " ").upper()}"')
        else:
            # Range match
            formula_parts.append(f'IF(AND({averagecell}>={min_val}, {averagecell}<={max_val}), "{str(rating_name).replace("_", " ").upper()}"')
    
    # Close all nested IFs
    formula = ""
    for i, part in enumerate(formula_parts):
        formula += part + ", "
    formula += '"UNRATED"' + ")" * len(formula_parts)


    wformula_parts = []
    for wrating_name, thresholds in rating_thresholds.items():
        min_val = thresholds.get("min", 0)
        max_val = thresholds.get("max", 5)
        if min_val == max_val:
            # Exact match (e.g., OUTSTANDING = 5)
            wformula_parts.append(f'IF({a["rating"]["weighted_avg"]}={min_val}, "{str(wrating_name).replace("_", " ").upper()}"')
        else:
            # Range match
            wformula_parts.append(f'IF(AND({a["rating"]["weighted_avg"]}>={min_val}, {a["rating"]["weighted_avg"]}<={max_val}), "{str(wrating_name).replace("_", " ").upper()}"')
    
    # Close all nested IFs
    wformula = ""
    for i, part in enumerate(wformula_parts):
        wformula += part + ", "
    wformula += '"UNRATED"' + ")" * len(wformula_parts)
    
    ws[f"N{row}"] = 0 if is_draft else f"={formula}"
    ws[f"N{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"N{row}"].font = Font(bold=True)

    ws[f"R{row}"] = 0 if is_draft else f"={wformula}"
    ws[f"R{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"R{row}"].font = Font(bold=True)

    # individuals block (Discussed, Assessed, Final, Confirm)
    row += 3
    prepareCells(ws, f"A{row}", f"D{row}")
    ws[f"A{row}"] = "Discussed With"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True)

    prepareCells(ws, f"E{row}", f"F{row}")
    ws[f"E{row}"] = "Date"
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Assessed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prepareCells(ws, f"M{row}", f"P{row}")
    ws[f"M{row}"] = "Final Rating By"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True)

    prepareCells(ws, f"Q{row}", f"S{row}")
    ws[f"Q{row}"] = "Date"
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    # name + position for discussed/assess/final
    prepareCells(ws, f"A{row}", f"D{row+2}")
    ws[f"A{row}"] = f"{individuals.get('discuss', {}).get('name','')}\n{individuals.get('discuss', {}).get('position','')}"
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"A{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"E{row}", f"F{row+2}")
    ws[f"E{row}"] = ""
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"E{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('assess', {}).get('name','')}\n{individuals.get('assess', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"M{row}", f"P{row+2}")
    ws[f"M{row}"] = f"{individuals.get('final', {}).get('name','')}\n{individuals.get('final', {}).get('position','')}"
    ws[f"M{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"M{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"Q{row}", f"S{row+2}")
    ws[f"Q{row}"] = ""
    ws[f"Q{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"Q{row}"].font = Font(bold=True, name="Calibri", size="11")

    row += 4
    prepareCells(ws, f"G{row}", f"J{row}")
    ws[f"G{row}"] = "Confirmed By"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True)

    prepareCells(ws, f"K{row}", f"L{row}")
    ws[f"K{row}"] = "Date"
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    prepareCells(ws, f"G{row}", f"J{row+2}")
    ws[f"G{row}"] = f"{individuals.get('confirm', {}).get('name','')}\n{individuals.get('confirm', {}).get('position','')}"
    ws[f"G{row}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws[f"G{row}"].font = Font(bold=True, name="Calibri", size="11")

    prepareCells(ws, f"K{row}", f"L{row+2}")
    ws[f"K{row}"] = ""
    ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[f"K{row}"].font = Font(bold=True, name="Calibri", size="11")

    # Save file
    id_rand = random.randint(1, 999999)
    fname_given = given
    fname_middle = middle[0] if middle else ""
    fname_last = last
    prefix = filename_prefix if filename_prefix else "IPCR"
    filename = f"{prefix} {period} - {fname_last} - {id_rand}"
    link = f"excels/IPCR/{filename}.xlsx"
    wb.save(link)

    file_url = upload_file(link, "commithub-bucket", f"IPCR/{filename}.xlsx")


    return file_url
