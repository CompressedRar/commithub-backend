# utils/DepartmentReportHandler.py
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import random
from utils.FileStorage import upload_file
from models.Departments import Department
from models.User import User
from models.System_Settings import System_Settings

def prepareCells(workshop, start, end):
    workshop.merge_cells(f"{start}:{end}")
    
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in workshop[f"{start}:{end}"]:
        for cell in row:
            cell.border = border_style

def create_department_performance_report(department_id, filename_prefix=None):
    """
    Create an Excel sheet showing average performance of members in a department.
    
    Layout:
    - Header: Office name | Performance Assessment: Rating
    - Columns: Name | Numerical | Adjective
    - Data rows with employee ratings
    - Footer: Total, No. of Employees, Average Ratings
    
    Args:
        department_id: The department DB id
        filename_prefix: Optional prefix for filename
    
    Returns:
        download link from FileStorage
    """
    
    # Query department and members
    dept = Department.query.get(department_id)
    if not dept:
        raise ValueError("Department not found")
    
    users = User.query.filter_by(department_id=department_id, account_status=1).all()
    
    # Get system settings for rating thresholds
    settings = System_Settings.get_default_settings()
    rating_thresholds = settings.rating_thresholds if settings else {
        "Outstanding": {"min": 4.5, "max": 5.0},
        "Very Satisfactory": {"min": 3.5, "max": 4.49},
        "Satisfactory": {"min": 2.5, "max": 3.49},
        "Unsatisfactory": {"min": 1.5, "max": 2.49},
        "Poor": {"min": 0, "max": 1.49}
    }
    
    # Create new workbook
    wb = load_workbook("excels/SummaryReportIPCRF.xlsx")
    
    ws = wb.active
    ws.title = "Department Performance"
    
    
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_font = Font(bold=True)
    
    # Row 1-3: Headers
    ws["G8"] = dept.name
    
    
    
    # Collect member data
    member_data = []
    total_rating = 0
    
    for user in users:
        # Calculate average performance for this user
        # Using IPCR final average if available, otherwise 0
        from models.PCR import IPCR
        active_ipcrs = IPCR.query.filter_by(
            user_id=user.id,
            status=1,
            period=settings.current_period_id if settings else None
        ).all()
        
        avg_rating = user.calculatePerformance()
        
        member_data.append({
                'name': f"{user.first_name} {user.middle_name[0] + '.' if user.middle_name else ''} {user.last_name}".strip(),
                'rating': round(avg_rating, 2)
            })
        total_rating += avg_rating
    
    # Sort by name

    print("TOTAL RATING:",total_rating)
    member_data.sort(key=lambda x: x['name'])
    
    # Write member data rows
    row = 10
    for member in member_data:
        prepareCells(ws, "G"+str(row), "I"+str(row))
        prepareCells(ws, "J"+str(row), "L"+str(row))
        prepareCells(ws, "M"+str(row), "O"+str(row))
        name = member['name']
        rating = member['rating']
        
        # Determine adjective rating
        adjective = "UNRATED"
        for rating_name, thresholds in rating_thresholds.items():
            min_val = thresholds.get("min", 0)
            max_val = thresholds.get("max", 5)
            if min_val <= rating <= max_val:
                adjective = str(rating_name).replace("_", " ").upper()
                break
        
        # Write row
        ws[f'G{row}'] = name
        ws[f'J{row}'] = rating
        ws[f'M{row}'] = adjective
        
        # Apply formatting
        for col in ['G', 'J', 'M']:
            cell = ws[f'{col}{row}']
            cell.border = border_style
            cell.alignment = Alignment(horizontal="left" if col == 'G' else "center", vertical="center")
            if col == 'J':
                cell.number_format = '0.00'
        
        row += 1
    
    # Empty row
    row += 1

    prepareCells(ws, "G"+str(row), "I"+str(row))
    prepareCells(ws, "J"+str(row), "L"+str(row))
    prepareCells(ws, "M"+str(row), "O"+str(row))
    
    # Total row
    ws[f'G{row}'] = "Total"
    ws[f'J{row}'] = round(total_rating, 2) if member_data else 0
    
    for col in ['G', 'J']:
        cell = ws[f'{col}{row}']
        cell.font = total_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col == 'J':
            cell.number_format = '0.00'
    
    row += 1

    prepareCells(ws, "G"+str(row), "I"+str(row))
    prepareCells(ws, "J"+str(row), "L"+str(row))
    prepareCells(ws, "M"+str(row), "O"+str(row))
    
    # No. of Employees row
    ws[f'G{row}'] = "No. of Employees"
    ws[f'J{row}'] = len(member_data)
    
    for col in ['G', 'J']:
        cell = ws[f'{col}{row}']
        cell.font = total_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    row += 1

    prepareCells(ws, "G"+str(row), "I"+str(row))
    prepareCells(ws, "J"+str(row), "L"+str(row))
    prepareCells(ws, "M"+str(row), "O"+str(row))
    
    # Average Ratings row
    avg_rating = total_rating / len(member_data) if member_data else 0
    
    # Determine average adjective
    avg_adjective = "UNRATED"
    for rating_name, thresholds in rating_thresholds.items():
        min_val = thresholds.get("min", 0)
        max_val = thresholds.get("max", 5)
        if min_val <= avg_rating <= max_val:
            avg_adjective = str(rating_name).replace("_", " ").upper()
            break
    
    ws[f'G{row}'] = "Average Ratings of Staff"
    ws[f'J{row}'] = round(avg_rating, 2)
    ws[f'M{row}'] = str(avg_adjective).replace("_", " ").upper()
    
    for col in ['G', 'J', 'M']:
        cell = ws[f'{col}{row}']
        cell.font = total_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col == 'J':
            cell.number_format = '0.00'
    
    # Save and upload
    id_rand = random.randint(1, 999999)
    prefix = filename_prefix if filename_prefix else "Department Performance"
    period = f"{datetime.datetime.now().strftime('%B %Y')}"
    filename = f"{prefix} - {dept.name} - {period} - {id_rand}"
    
    filepath = f"excels/DepartmentReports/{filename}.xlsx"
    wb.save(filepath)
    
    # Upload to cloud storage
    file_url = upload_file(filepath, "commithub-bucket", f"DepartmentReports/{filename}.xlsx")
    
    return file_url
